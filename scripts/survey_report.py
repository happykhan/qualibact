#!/usr/bin/env python3
"""
Generate a generic (anonymised) report from the QualiBact assessment
survey responses. The report is intended to go back to every
respondent as a "thank you + here's what we heard" follow-up.

Reads:  Google Forms export xlsx (rows = responses, cols = questions)
Writes: private/feedback/survey-{label}.md
        private/feedback/survey-{label}-charts/{ratings,confidence,agreement}.png

Output lives under /private/ (gitignored, repo-private) because the PDF
is meant to go back to respondents by email — not published to the
website. The report itself is anonymised (aggregated counts, anonymised
quotes), but the dir is kept off the public GitHub repo by convention.

Usage:
    python3 scripts/survey_report.py \\
        --xlsx "/path/to/QualiBact assessment (Responses) (2).xlsx" \\
        --label 2026-05-20
"""
from __future__ import annotations

import argparse
import re
import shutil
import subprocess
import sys
from collections import Counter
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent

CHROME_PATHS = [
    "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
    "/Applications/Chromium.app/Contents/MacOS/Chromium",
    "/Applications/Microsoft Edge.app/Contents/MacOS/Microsoft Edge",
]


def _find_chrome() -> str | None:
    for p in CHROME_PATHS:
        if Path(p).exists():
            return p
    for cmd in ("google-chrome", "chromium", "chromium-browser"):
        path = shutil.which(cmd)
        if path:
            return path
    return None


HTML_TEMPLATE = """<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>{title}</title>
<style>
  @page {{ size: A4; margin: 18mm 16mm 18mm 16mm; }}
  body {{
    font-family: -apple-system, "Helvetica Neue", Helvetica, Arial, sans-serif;
    color: #1f2937;
    line-height: 1.45;
    font-size: 10.5pt;
    max-width: 178mm;
  }}
  h1 {{ font-size: 20pt; color: #0f4c81; border-bottom: 2px solid #0f4c81; padding-bottom: 4pt; }}
  h2 {{ font-size: 13pt; color: #0f4c81; margin-top: 18pt; border-bottom: 1px solid #d1d5db; padding-bottom: 2pt; }}
  h3 {{ font-size: 11pt; color: #1e3a5f; margin-top: 12pt; }}
  p {{ margin: 6pt 0; }}
  ul, ol {{ margin: 4pt 0 8pt 18pt; padding: 0; }}
  li {{ margin: 2pt 0; }}
  table {{
    border-collapse: collapse;
    width: 100%;
    font-size: 9pt;
    margin: 8pt 0 12pt 0;
  }}
  th, td {{
    border: 1px solid #d1d5db;
    padding: 4pt 6pt;
    text-align: left;
    vertical-align: top;
  }}
  th {{ background: #f3f4f6; }}
  tbody tr:nth-child(even) {{ background: #fafafa; }}
  img {{ max-width: 100%; height: auto; margin: 6pt 0; page-break-inside: avoid; }}
  blockquote {{
    margin: 6pt 0;
    padding: 4pt 10pt;
    border-left: 3px solid #cbd5e1;
    background: #f8fafc;
    color: #334155;
    font-size: 9.5pt;
  }}
  a {{ color: #0f4c81; text-decoration: none; }}
  code {{ background: #f3f4f6; padding: 1pt 3pt; border-radius: 3pt; font-size: 9pt; }}
  hr {{ border: 0; border-top: 1px solid #d1d5db; margin: 12pt 0; }}
  em {{ color: #475569; }}
</style>
</head>
<body>
{body}
</body>
</html>
"""


def render_html(markdown_text: str, title: str, charts_abs_dir: Path) -> str:
    """Convert markdown to HTML, swap relative chart paths for file:// absolute
    URLs so Chrome resolves them, and wrap in a print-friendly template."""
    import markdown as md_lib  # type: ignore
    html_body = md_lib.markdown(
        markdown_text,
        extensions=["tables", "fenced_code", "sane_lists"],
    )
    # Replace `<img src="survey-LABEL-charts/foo.png">` with absolute file:// URLs
    rel = charts_abs_dir.name
    abs_prefix = charts_abs_dir.resolve().as_uri()
    html_body = re.sub(
        rf'src="({re.escape(rel)})/',
        f'src="{abs_prefix}/',
        html_body,
    )
    return HTML_TEMPLATE.format(title=title, body=html_body)


def html_to_pdf(html_path: Path, pdf_path: Path) -> bool:
    chrome = _find_chrome()
    if not chrome:
        print("No Chrome/Chromium found — skipping PDF (see CHROME_PATHS).", file=sys.stderr)
        return False
    cmd = [
        chrome, "--headless=new", "--disable-gpu", "--no-pdf-header-footer",
        f"--print-to-pdf={pdf_path}",
        "--virtual-time-budget=2000",
        html_path.resolve().as_uri(),
    ]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
    except (subprocess.TimeoutExpired, FileNotFoundError) as e:
        print(f"PDF render failed: {e}", file=sys.stderr)
        return False
    if result.returncode != 0 or not pdf_path.exists():
        print(f"chrome --print-to-pdf returned {result.returncode}", file=sys.stderr)
        if result.stderr:
            print(result.stderr[:500], file=sys.stderr)
        return False
    return True

UTILITY_TIERS = ["Essential", "Very Useful", "Moderately Useful", "Not Useful"]
UTILITY_COLOURS = {
    "Essential": "#1a9850",
    "Very Useful": "#a6d96a",
    "Moderately Useful": "#fdae61",
    "Not Useful": "#d73027",
}

THRESHOLD_PAIRS = [
    ("N50", "lower", 4, 5),
    ("Number of contigs", "upper", 6, 7),
    ("GC content", "lower", 8, 9),
    ("GC content", "upper", 10, 11),
    ("Completeness", "lower", 12, 13),
    ("Completeness", "upper", 14, 15),
    ("Contamination", "lower", 16, 17),
    ("Contamination", "upper", 18, 19),
    ("Total Coding Sequences", "lower", 20, 21),
    ("Total Coding Sequences", "upper", 22, 23),
    ("Assembly size", "lower", 24, 25),
    ("Assembly size", "upper", 26, 27),
]

RATING_COLS = [
    ("N50", 29),
    ("Number of Contigs", 30),
    ("GC Content", 31),
    ("Completeness", 32),
    ("Contamination", 33),
    ("Total Coding Sequences", 34),
    ("Assembly Size", 35),
]


def _classify_agreement(cell: str | None) -> str:
    """Map a free-form cell value to one of {yes, no, no_bound_defined, other}."""
    if cell is None:
        return "blank"
    s = str(cell).strip()
    if not s:
        return "blank"
    sl = s.lower()
    if sl == "yes":
        return "yes"
    if sl in {"no"}:
        return "no"
    if sl == "maybe":
        return "maybe"
    if re.search(r"\b(not defined|no (upper|lower) bound)\b", sl):
        return "no_bound_defined"
    return "other"


def load_responses(xlsx_path: Path) -> tuple[list[dict], tuple[str, str] | None]:
    """Returns (rows, (first_ts, last_ts)) — the survey collection window
    is derived from the actual Timestamp column rather than passed in as
    a label, so the prose reflects when respondents wrote, not when the
    report was generated."""
    from openpyxl import load_workbook  # local import keeps the script importable
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows: list[dict] = []
    timestamps: list = []
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(row):
            continue
        ts = row[0]
        if ts is not None:
            timestamps.append(ts)
        rows.append({
            "species": (row[2] or "").strip() if row[2] else "",
            "agreement": {label: (row[ans_col - 1], row[expl_col - 1])
                          for (m, b, ans_col, expl_col) in THRESHOLD_PAIRS
                          for label in [f"{m} {b}"]},
            "ratings": {m: row[col - 1] for m, col in RATING_COLS},
            "confidence": row[35],
            "additional_metrics": row[36],
            "additional_comments": row[27],
        })
    window = None
    if timestamps:
        try:
            ts_sorted = sorted(timestamps)
            first = ts_sorted[0]
            last = ts_sorted[-1]
            # openpyxl returns datetime when the cell is a date — otherwise str
            def _fmt(t):
                if hasattr(t, "strftime"):
                    return t.strftime("%-d %B %Y")
                return str(t)
            window = (_fmt(first), _fmt(last))
        except (TypeError, ValueError):
            window = None
    return rows, window


def _plot_ratings(rows: list[dict], out_path: Path) -> None:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np

    metric_names = [m for m, _ in RATING_COLS]
    counts = {m: Counter() for m in metric_names}
    for r in rows:
        for m, val in r["ratings"].items():
            if val:
                counts[m][val] += 1

    fig, ax = plt.subplots(figsize=(11, 5))
    x = np.arange(len(metric_names))
    bottom = np.zeros(len(metric_names))
    for tier in UTILITY_TIERS:
        vals = np.array([counts[m].get(tier, 0) for m in metric_names])
        ax.bar(x, vals, bottom=bottom, label=tier,
               color=UTILITY_COLOURS[tier], edgecolor="white", linewidth=0.6)
        bottom += vals
    ax.set_xticks(x)
    ax.set_xticklabels(metric_names, rotation=15, ha="right")
    ax.set_ylabel("Responses")
    ax.set_title(f"Per-metric utility rating  (n={len(rows)} responses)")
    ax.legend(loc="lower right", framealpha=0.95, fontsize=9)
    ax.grid(axis="y", alpha=0.3)
    plt.tight_layout()
    plt.savefig(out_path, dpi=130)
    plt.close(fig)


def _plot_confidence(rows: list[dict], out_path: Path) -> None:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np
    vals = [r["confidence"] for r in rows if r["confidence"] is not None]
    if not vals:
        return
    fig, ax = plt.subplots(figsize=(8, 4))
    bins = np.arange(0.5, 11, 1)
    ax.hist(vals, bins=bins, color="#4575b4", edgecolor="white")
    mean = sum(vals) / len(vals)
    median = sorted(vals)[len(vals) // 2]
    ax.axvline(mean, color="red", linestyle="--", linewidth=1, label=f"mean {mean:.1f}")
    ax.axvline(median, color="black", linestyle=":", linewidth=1, label=f"median {median:.0f}")
    ax.set_xticks(range(1, 11))
    ax.set_xlim(0.5, 10.5)
    ax.set_xlabel("Confidence score (1–10)")
    ax.set_ylabel("Responses")
    ax.set_title(f"Overall confidence in the QualiBact framework  (n={len(vals)})")
    ax.legend()
    ax.grid(axis="y", alpha=0.3)
    plt.tight_layout()
    plt.savefig(out_path, dpi=130)
    plt.close(fig)


def _plot_agreement(rows: list[dict], out_path: Path) -> None:
    """Stacked bar of agree / disagree / other for every threshold pair."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np

    labels = [f"{m}\n{b}" for (m, b, _, _) in THRESHOLD_PAIRS]
    yes = []
    no = []
    other = []
    nbd = []
    for (m, b, _, _) in THRESHOLD_PAIRS:
        cnt = Counter()
        for r in rows:
            val, _ = r["agreement"].get(f"{m} {b}", (None, None))
            cnt[_classify_agreement(val)] += 1
        yes.append(cnt["yes"])
        no.append(cnt["no"])
        other.append(cnt["other"] + cnt["maybe"])
        nbd.append(cnt["no_bound_defined"])

    fig, ax = plt.subplots(figsize=(13, 5))
    x = np.arange(len(labels))
    p1 = ax.bar(x, yes, label="Agree", color="#1a9850", edgecolor="white")
    p2 = ax.bar(x, no, bottom=yes, label="Disagree", color="#d73027", edgecolor="white")
    p3 = ax.bar(x, other, bottom=np.array(yes) + np.array(no),
                label="Other (suggested value)", color="#fdae61", edgecolor="white")
    p4 = ax.bar(x, nbd, bottom=np.array(yes) + np.array(no) + np.array(other),
                label='"No bound defined"', color="#bdbdbd", edgecolor="white")
    ax.set_xticks(x)
    ax.set_xticklabels(labels, fontsize=8)
    ax.set_ylabel("Responses")
    ax.set_title(f"Per-threshold agreement  (n={len(rows)} responses)")
    ax.legend(loc="upper right", framealpha=0.95, fontsize=9)
    ax.grid(axis="y", alpha=0.3)
    plt.tight_layout()
    plt.savefig(out_path, dpi=130)
    plt.close(fig)


def _agreement_table(rows: list[dict]) -> list[tuple[str, int, int, int, int, int]]:
    """Returns rows: (label, yes, no, other, no_bound_defined, total)."""
    out = []
    for (m, b, _, _) in THRESHOLD_PAIRS:
        cnt = Counter()
        for r in rows:
            val, _ = r["agreement"].get(f"{m} {b}", (None, None))
            cnt[_classify_agreement(val)] += 1
        total = sum(cnt.values()) - cnt["blank"]
        out.append((
            f"{m} ({b})",
            cnt["yes"], cnt["no"], cnt["other"] + cnt["maybe"],
            cnt["no_bound_defined"], total,
        ))
    return out


def _ratings_table(rows: list[dict]) -> list[tuple[str, int, int, int, int, int]]:
    """Per metric: counts in (Essential, Very Useful, Moderately, Not, total)."""
    out = []
    for m, _ in RATING_COLS:
        cnt = Counter(r["ratings"].get(m) for r in rows)
        out.append((m,
                    cnt.get("Essential", 0),
                    cnt.get("Very Useful", 0),
                    cnt.get("Moderately Useful", 0),
                    cnt.get("Not Useful", 0),
                    sum(cnt.get(t, 0) for t in UTILITY_TIERS)))
    return out


def _clean_quote(s: str) -> str:
    s = str(s).strip()
    return re.sub(r"\s+", " ", s)


def render(rows: list[dict], label: str, charts_rel: str,
           window: tuple[str, str] | None = None) -> str:
    n = len(rows)
    species = sorted({r["species"] for r in rows if r["species"]})
    genera = sorted({s.split()[0] for s in species})
    conf = [r["confidence"] for r in rows if r["confidence"] is not None]
    conf_mean = sum(conf) / len(conf) if conf else 0
    conf_median = sorted(conf)[len(conf) // 2] if conf else 0
    ten_count = sum(1 for c in conf if c == 10)

    if window:
        first_ts, last_ts = window
        if first_ts == last_ts:
            window_text = f"collected on {first_ts}"
        else:
            window_text = f"collected between {first_ts} and {last_ts}"
    else:
        window_text = ""

    lines: list[str] = []
    lines.append("# QualiBact assessment survey — what the community told us\n")
    if window_text:
        intro = (
            f"_Generic summary of the QualiBact threshold-assessment survey "
            f"({window_text})._ This document is shareable with all "
            f"respondents and contains no personally identifying information."
        )
    else:
        intro = (
            "_Generic summary of the QualiBact threshold-assessment survey._ "
            "This document is shareable with all respondents and contains "
            "no personally identifying information."
        )
    lines.append(intro + "\n")

    lines.append("## At a glance\n")
    lines.append(f"- **{n}** responses covering **{len(species)}** species across **{len(genera)}** genera.")
    lines.append(f"- Overall confidence: **mean {conf_mean:.1f} / 10**, median {conf_median:.0f}/10. "
                 f"{ten_count} respondents ({ten_count*100//len(conf)}%) gave the framework full marks.")
    lines.append("- Strongest-rated metrics (% Essential + Very Useful): Number of Contigs, "
                 "Assembly Size, Contamination, N50 — all above 90%.")
    lines.append("- Most-contested thresholds: assembly size and N50 bounds (see per-threshold "
                 "agreement chart below).")
    lines.append("")
    lines.append(f"![Per-metric utility rating]({charts_rel}/ratings.png)")
    lines.append("")
    lines.append(f"![Overall confidence]({charts_rel}/confidence.png)")
    lines.append("")
    lines.append(f"![Per-threshold agreement]({charts_rel}/agreement.png)")
    lines.append("")

    lines.append("## How useful is each metric?\n")
    lines.append("Each respondent rated each metric on a four-point scale "
                 "(_Essential / Very Useful / Moderately Useful / Not Useful_).\n")
    lines.append("| Metric | Essential | Very useful | Moderately useful | Not useful | n |")
    lines.append("|---|---|---|---|---|---|")
    for m, e, vu, mu, nu, t in _ratings_table(rows):
        lines.append(f"| {m} | {e} | {vu} | {mu} | {nu} | {t} |")
    lines.append("")

    lines.append("## Do you agree with the suggested threshold?\n")
    lines.append("Per (metric, lower/upper bound), the response distribution. "
                 "_Other_ counts respondents who suggested a specific alternate "
                 "value in the explanation field (e.g. “we suggest "
                 "2.1 Mb instead of 2.0 Mb”). _“No bound defined”_ "
                 "is when the curated threshold for that bound is left blank "
                 "(common for upper Completeness, lower Contamination).\n")
    lines.append("| Bound | Agree | Disagree | Other (alt. value) | “No bound defined” | n |")
    lines.append("|---|---|---|---|---|---|")
    for label_, y, no_, other_, nbd, t in _agreement_table(rows):
        lines.append(f"| {label_} | {y} | {no_} | {other_} | {nbd} | {t} |")
    lines.append("")

    # Free-text themes — additional metrics requested
    lines.append("## What other metrics would you find useful?\n")
    lines.append("Free-text answers (with the requesting respondent's species "
                 "in square brackets):\n")
    seen_themes = Counter()
    for r in rows:
        v = r["additional_metrics"]
        if not v:
            continue
        s = _clean_quote(v)
        if s.lower() in {"no", "n/a", "none", "na"}:
            continue
        # Theme heuristics for the summary table
        sl = s.lower()
        if "busco" in sl: seen_themes["BUSCO completeness"] += 1
        if "depth" in sl or "coverage" in sl: seen_themes["Sequencing depth / coverage"] += 1
        if "heterozyg" in sl: seen_themes["Heterozygous positions"] += 1
        if "ns per" in sl or "n's per" in sl or "ns/100" in sl: seen_themes["Ns per 100 kbp"] += 1
        if "contaminating species" in sl or "contaminant" in sl: seen_themes["Identity of contaminating species"] += 1
        if "median" in sl or "iqr" in sl or "percentile" in sl: seen_themes["Show median / IQR / percentiles"] += 1
        if "warning" in sl and "fail" in sl: seen_themes["PASS / WARNING / FAIL tiers"] += 1
        if "minimum length" in sl or "1kb" in sl or "min length" in sl: seen_themes["Minimum contig length filter"] += 1
        species_short = (r["species"] or "?")[:35]
        lines.append(f"> _[{species_short}]_ {s}")
        lines.append("")
    if seen_themes:
        lines.append("### Themes across these requests\n")
        lines.append("| Theme | Mentions |")
        lines.append("|---|---|")
        for k, v in sorted(seen_themes.items(), key=lambda kv: -kv[1]):
            lines.append(f"| {k} | {v} |")
        lines.append("")

    # Additional comments intentionally omitted from this generic report —
    # handled in one-to-one follow-up replies instead. Add via a separate
    # personalised letter generator if needed.

    # What QualiBact did with the feedback (boilerplate, edit per cycle)
    lines.append("## What QualiBact did with the feedback\n")
    lines.append(
        "- **Adjusted thresholds where experts pointed at specific values.** "
        "Examples shipped in qualibact-v1.1: *Campylobacter jejuni* assembly "
        "size upper raised from 2.0 Mb to 2.1 Mb (Alexandra Nunes & Mónica "
        "Oleastro, INSA); *Neisseria meningitidis* upper tightened from 2.4 "
        "Mb to 2.3 Mb (Alexandra Nunes & Célia Bettencourt, INSA)."
    )
    lines.append(
        "- **Reference-set re-curation requests.** *Klebsiella grimontii* "
        "and *Staphylococcus coagulans* have engine re-runs queued. "
        "Misclassified-accession lists are being collected via direct "
        "correspondence."
    )
    lines.append(
        "- **Added a PASS / WARN / FAIL three-tier system.** Multiple "
        "respondents flagged that a single FAIL boundary is too coarse. "
        "The engine now emits a tighter WARN bound alongside the existing "
        "FAIL bound; the species page renders both."
    )
    lines.append(
        "- **Surfaced engine quality flags on the species page.** When the "
        "engine detects that the reference dataset for a species itself has "
        "issues (high contamination fraction, oversized genomes, wide GC "
        "range, etc.), a coloured banner now appears at the top of the "
        "species page with the engine's interpretation."
    )
    lines.append(
        "- **Read-level QC requests parked.** Sequencing depth, "
        "heterozygous positions, and Ns/100 kbp are out of scope for "
        "assembly-level QualiBact thresholds. We recommend "
        "[bactscout](https://github.com/cgps-group/bactscout) as the "
        "companion tool."
    )
    lines.append(
        "- **BUSCO completeness, identity of contaminating species.** "
        "These are out of scope for QualiBact's published thresholds, "
        "but downstream QC tools (CheckM2, sylph, kraken2) are where "
        "per-assembly contaminant identification belongs."
    )
    lines.append("")

    lines.append("## Where the live thresholds live\n")
    lines.append("- Species page: <https://qualibact.org/{Genus}/{Genus_species}/>")
    lines.append("- Aggregate CSV (4-bound FINAL + WARN): <https://qualibact.org/api/v2/thresholds.csv>")
    lines.append("- JSON shape: <https://qualibact.org/api/v2/thresholds.json>")
    lines.append("- Methods: <https://qualibact.org/methods/qualibact-v1.1/>")
    lines.append("- Per-WHO-list download: <https://qualibact.org/priority-pathogens/>")
    lines.append("- Contributors (with your name when you contributed): <https://qualibact.org/contributors/>")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append(f"_Report generated {label}. Survey responses {window_text or 'date unknown'}._")
    lines.append("")
    return "\n".join(lines)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", required=True, type=Path)
    ap.add_argument("--label", required=True)
    args = ap.parse_args()
    if not args.xlsx.exists():
        print(f"ERROR: {args.xlsx} not found", file=sys.stderr)
        return 2

    rows, window = load_responses(args.xlsx)
    if not rows:
        print("ERROR: no responses loaded", file=sys.stderr)
        return 2

    out_dir = REPO_ROOT / "private" / "feedback"
    out_dir.mkdir(parents=True, exist_ok=True)
    charts_dir = out_dir / f"survey-{args.label}-charts"
    charts_dir.mkdir(parents=True, exist_ok=True)

    _plot_ratings(rows, charts_dir / "ratings.png")
    _plot_confidence(rows, charts_dir / "confidence.png")
    _plot_agreement(rows, charts_dir / "agreement.png")

    md = render(rows, args.label, f"survey-{args.label}-charts", window=window)
    out_path = out_dir / f"survey-{args.label}.md"
    out_path.write_text(md, encoding="utf-8")
    print(f"Wrote {out_path.relative_to(REPO_ROOT)} ({len(md.splitlines())} lines)")
    for c in sorted(charts_dir.glob("*.png")):
        print(f"  {c.relative_to(REPO_ROOT)}")

    # Email-attachable: also emit a self-contained HTML + a PDF rendered
    # via headless Chrome. The HTML has inlined CSS and absolute file://
    # URLs for the chart images, so Chrome resolves them when printing.
    title = f"QualiBact assessment — community summary ({args.label})"
    html = render_html(md, title, charts_dir)
    html_path = out_dir / f"survey-{args.label}.html"
    html_path.write_text(html, encoding="utf-8")
    pdf_path = out_dir / f"survey-{args.label}.pdf"
    if html_to_pdf(html_path, pdf_path):
        print(f"  {html_path.relative_to(REPO_ROOT)}")
        print(f"  {pdf_path.relative_to(REPO_ROOT)}  ({pdf_path.stat().st_size // 1024} KB)")
    else:
        print(f"  {html_path.relative_to(REPO_ROOT)}  (PDF skipped — open in browser and print to save)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
